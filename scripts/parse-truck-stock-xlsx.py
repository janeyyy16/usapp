"""
Parse the company-wide inventory workbook and extract the Truck Stock
section of each branch sheet into a flat JSON file that the Truck Stock
page can bulk-import into Supabase.

The workbook layout is column-based per sheet: rows 1-2 are header
labels and row 3 is blank; each sheet has multiple sections sitting
side-by-side (RETURN, LIVE, TS / TRUCK STOCK / TRUCK STOP). We
auto-detect the TS section by scanning row 2 for one of those tokens,
then read PART #, Location, Status, Notes from that section onwards.

Branch name = sheet name (the user knows their internal abbreviations).
Sheets that are non-branch (FORM, NEED RETURN, BO SHIPPED, BLANK …)
are skipped.

Output:
  src/lib/truckStockSeed.json  -> array of { branch, partNo, description, manufacturer, quantity, notes }
"""
import json
import re
from pathlib import Path
import openpyxl

SRC = Path(r"C:\\Users\\user\\Downloads\\PROJECT A USAPP ALL PARTS INVENTORY USE.xlsx")
OUT = Path(__file__).resolve().parent.parent / "src" / "lib" / "truckStockSeed.json"

SKIP_SHEETS = {"FORM", "Sheet44", "Copy of RL", "Sheet17", "NEED RETURN", "BO SHIPPED", "BLANK (DONT EDIT)"}

# Tokens that mark a Truck Stock section header in row 2 of each sheet.
TS_TOKENS = ("TS", "TRUCK STOCK", "TRUCK STOP", "TRUCK-STOCK")


PART_HEADERS = {"PART", "PART #", "PART NO", "PART NUMBER", "PARTNUMBER", "PARTNO", "PART#"}
LOCATION_HEADERS = {"LOCATION", "BIN", "SHELF"}
STATUS_HEADERS = {"STATUS", "CONDITION", "STATE"}
QUANTITY_HEADERS = {"QUANTITY", "QTY", "STOCK"}
NOTE_HEADERS = {"NOTE", "NOTES", "COMMENT", "COMMENTS", "RELATION", "TRACKING", "TICKET", "TICKET #",
                "DATE PICKED", "TECH", "VENDOR", "AGING?", "CX NAME"}


def header_kind(value):
    """Classify a row-1 header cell so we can map columns by meaning."""
    if value is None:
        return None
    v = str(value).strip().upper().replace("_", " ").replace(".", "")
    if not v:
        return None
    if v in PART_HEADERS:
        return "part"
    if v in LOCATION_HEADERS:
        return "location"
    if v in STATUS_HEADERS:
        return "status"
    if v in QUANTITY_HEADERS:
        return "quantity"
    if v in NOTE_HEADERS:
        return "note"
    return None


def resolve_section_columns(header_row1, part_col):
    """Walk outward from the PART # column on row 1 and bind each known
    field name to its actual column index. We stop at the next PART #
    (so we don't bleed into an adjacent section)."""
    bindings = {"part": part_col}
    # Walk RIGHT.
    for j in range(part_col + 1, min(part_col + 8, len(header_row1))):
        kind = header_kind(header_row1[j])
        if kind == "part":
            break  # next section starts here
        if kind and kind not in bindings:
            bindings[kind] = j
    # Walk LEFT (handles CG-style sheets where Location/Status sit to
    # the left of Part #). Stop at the previous section's last column.
    for j in range(part_col - 1, max(part_col - 6, -1), -1):
        kind = header_kind(header_row1[j])
        if kind == "part":
            break
        if kind and kind not in bindings:
            bindings[kind] = j
    return bindings


def is_ts_marker(value):
    if value is None:
        return False
    v = str(value).strip().upper().replace("  ", " ")
    if v in {"TS", "TRUCK STOCK", "TRUCK STOP", "TRUCK-STOCK"}:
        return True
    # Some sheets have "TS UPDATED 6/8/2026" etc. — still a TS section.
    return v.startswith("TS ") or "TRUCK STOCK" in v or "TRUCK STOP" in v


def looks_like_part_header(value):
    if value is None:
        return False
    v = str(value).strip().upper().replace(".", "").replace("_", "").replace("-", " ")
    return v in PART_HEADERS


def find_section_columns(header_row1, header_row2):
    """Return list of (part_col, label) tuples.

    For each TS marker in row 2 we walk row 1 forwards (and a couple of
    columns backwards as a fallback) looking for the next "PART #"
    header — that's the actual column we should read part numbers from.
    Some sheets put the marker in a different column from the PART #
    header (e.g. CG: marker @ 13, PART # @ 15).
    """
    sections = []
    for col_idx, marker in enumerate(header_row2):
        if not is_ts_marker(marker):
            continue
        part_col = None
        # 1. Walk forward in row 1 from the marker column, but stop if
        #    we run into another TS / non-TS marker further right.
        for j in range(col_idx, min(col_idx + 8, len(header_row1))):
            if looks_like_part_header(header_row1[j]):
                part_col = j
                break
        # 2. If we still haven't found one, fall back to a tight window
        #    around the marker column. Some sheets (LR, NF) put the
        #    PART # right under the TS marker.
        if part_col is None:
            for j in range(max(0, col_idx - 1), col_idx + 2):
                if j < len(header_row1) and looks_like_part_header(header_row1[j]):
                    part_col = j
                    break
        # 3. Last resort: use the marker column itself.
        if part_col is None:
            part_col = col_idx
        sections.append(part_col)
    # De-duplicate (same part column reachable from multiple markers).
    return sorted(set(sections))


def normalize_part_no(value):
    if value is None:
        return ""
    s = str(value).strip()
    # Excel often stores numeric part numbers as floats with .0 trailing.
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    # Some sheets prefix with apostrophe to keep leading zeros — drop it.
    if s.startswith("'"):
        s = s[1:]
    return s


def normalize_quantity(value):
    if value is None:
        return 1
    s = str(value).strip().lower()
    if not s:
        return 1
    # Common patterns: "x1", "x 2", "1.0", "2", "3 pcs"
    m = re.search(r"\d+(?:\.\d+)?", s)
    if not m:
        return 1
    n = float(m.group())
    return max(1, int(n))


def parse_sheet(ws, branch_name):
    """Walk the truck-stock columns of the sheet and yield row dicts.

    Each TS section's column layout differs per sheet — some have
    Location to the right of Part #, others (CG) have it to the left
    and put Quantity to the right. We bind columns by their row-1
    header label rather than fixed offsets.
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4:
        return []
    header2 = rows[1]
    part_cols = find_section_columns(rows[0], header2)
    if not part_cols:
        return []

    out = []
    for part_col in part_cols:
        bindings = resolve_section_columns(rows[0], part_col)
        loc_col   = bindings.get("location")
        stat_col  = bindings.get("status")
        qty_col   = bindings.get("quantity")
        note_col  = bindings.get("note")

        for r_idx in range(3, len(rows)):
            row = rows[r_idx]

            def get(c):
                if c is None or c >= len(row):
                    return None
                return row[c]

            part_no = normalize_part_no(get(part_col))
            if not part_no or part_no.upper() in PART_HEADERS:
                continue
            # Skip stray section markers.
            if part_no.upper() in {"RETURN", "RETURNS", "RESTOCK", "LIVE", "LIVE PART(S)",
                                   "TS", "TRUCK STOCK", "TRUCK STOP"}:
                continue

            storage_location = (str(get(loc_col)).strip() if get(loc_col) else "")
            status_raw = (str(get(stat_col)).strip() if get(stat_col) else "")
            note_raw = (str(get(note_col)).strip() if get(note_col) else "")

            # Quantity priority: explicit Quantity column > "x N" tokens
            # in note/status > default 1. We never sum numeric junk from
            # unknown columns.
            qty = 1
            qty_value = get(qty_col) if qty_col is not None else None
            if qty_value is not None:
                # Cells like "x2", "x 3", "2", "2.0" all parse the same way.
                m = re.search(r"\d+(?:\.\d+)?", str(qty_value))
                if m:
                    n = float(m.group())
                    if n >= 1:
                        qty = max(1, int(n))
            for c in (note_raw, status_raw):
                m = re.match(r"^[xX]\s*(\d+)$", c.strip())
                if m:
                    qty = max(qty, int(m.group(1)))

            note_bits = [b for b in (status_raw, note_raw) if b and not re.fullmatch(r"[xX]\s*\d+", b)]
            notes = " | ".join(note_bits)[:500]

            out.append({
                "branch": branch_name,
                "partNo": part_no,
                "description": "",
                "manufacturer": "",
                "quantity": qty,
                "storageLocation": storage_location[:120],
                "notes": notes,
            })
    return out


def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    all_rows = []
    per_branch = {}
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        rows = parse_sheet(ws, sheet_name)
        if rows:
            per_branch[sheet_name] = len(rows)
            all_rows.extend(rows)

    # De-duplicate within (branch, partNo) — bump quantity for repeats.
    dedup = {}
    for r in all_rows:
        key = (r["branch"], r["partNo"].lower())
        if key in dedup:
            dedup[key]["quantity"] += r["quantity"]
            # Keep the longer storage location string if a later row has more detail.
            if len(r.get("storageLocation") or "") > len(dedup[key].get("storageLocation") or ""):
                dedup[key]["storageLocation"] = r["storageLocation"]
            if len(r["notes"]) > len(dedup[key]["notes"]):
                dedup[key]["notes"] = r["notes"]
        else:
            dedup[key] = r
    merged = list(dedup.values())
    merged.sort(key=lambda r: (r["branch"], r["partNo"]))

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(merged, indent=2), encoding="utf-8")
    print(f"Wrote {OUT}")
    print(f"Total rows: {len(merged)}")
    print(f"Per-branch counts:")
    for b, n in sorted(per_branch.items()):
        print(f"  {b:6} {n}")


if __name__ == "__main__":
    main()
