"""Walk every sheet and surface tech names / cities visible in the LIVE
section so we can confidently map the short codes to full branch names."""
import openpyxl

PATH = r"C:\Users\user\Downloads\PROJECT A USAPP ALL PARTS INVENTORY USE.xlsx"
wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)

SKIP = {"FORM", "Sheet44", "Copy of RL", "Sheet17", "NEED RETURN", "BO SHIPPED", "BLANK (DONT EDIT)"}

# Tech names to look for so we can match to locations.ts:
LOOKUP = {
    # Atlanta
    "abel severino": "Atlanta", "abraham im": "Atlanta", "gerrell berg": "Atlanta",
    "jordan brown": "Atlanta", "joshua silva": "Atlanta",
    "kevin khaiphanliane": "Atlanta", "nathan napora": "Atlanta",
    # Birmingham
    "david sims": "Birmingham", "kenny shin": "Birmingham", "zonate grant": "Birmingham",
    # Cape Girardeau
    "alaska olinger": "Cape Girardeau", "deprece harris": "Cape Girardeau",
    "matthew nichols": "Cape Girardeau",
    # Chattanooga
    "austin ferguson": "Chattanooga", "christian andrews": "Chattanooga",
    "seven grinis": "Chattanooga",
    # Columbus
    "a'dejaun tyson": "Columbus", "matt simmons": "Columbus", "percy smith": "Columbus",
    # Dallas
    "lashamus dowell": "Dallas",
    # Destin
    "garrett mccarley": "Destin",
    # Huntsville
    "dylan lano": "Huntsville", "jordan stanley": "Huntsville",
    "nathan wagner": "Huntsville",
    # Jackson, MS
    "anthony leonard cavett": "Jackson, MS", "antonio smith": "Jackson, MS",
    "mikkel brown": "Jackson, MS", "terry davis": "Jackson, MS",
    "tywon ross": "Jackson, MS",
    # Jackson, TN
    "brandon phillips": "Jackson, TN", "christian clark": "Jackson, TN",
    "gabriel talley": "Jackson, TN", "jaylon yarbrough": "Jackson, TN",
    "justin parker": "Jackson, TN",
    # Jacksonville
    "bradley hollowell": "Jacksonville", "zakarya moradi": "Jacksonville",
    # Jonesboro
    "jason bateman": "Jonesboro",
    # Knoxville
    "alex myles": "Knoxville", "joshua rhinehart": "Knoxville",
    "zac coisman": "Knoxville",
    # Lake Charles
    "danny thornton": "Lake Charles",
    # Little Rock
    "andre riddle": "Little Rock", "darius brown": "Little Rock",
    "jonathan knox": "Little Rock", "nocona detten": "Little Rock",
    # Memphis
    "darrin stewart": "Memphis", "darryel burdette": "Memphis",
    "jeff lucas": "Memphis", "rico shaw": "Memphis",
    "sean smith": "Memphis",
    # Mobile
    "dominic holman": "Mobile", "jonathan colquett": "Mobile",
    "jonathon allen": "Mobile", "thaddaeus springfield": "Mobile",
    # Montgomery
    "andy oh": "Montgomery",
    # Nashville
    "baolin henry zhang": "Nashville", "john godfrey": "Nashville",
    "justin robertson": "Nashville", "leo sun": "Nashville",
    "steven kurvink": "Nashville",
    # New Orleans
    "cole mushinsky": "New Orleans", "cooper shaffett": "New Orleans",
    "corey cage": "New Orleans", "joseph wease": "New Orleans",
    "kurt merckel": "New Orleans", "ryder tourere": "New Orleans",
    # Norfolk
    "chris simpson": "Norfolk", "edward lindsey": "Norfolk",
    # Raleigh
    "alexxis henry": "Raleigh", "damon ottley": "Raleigh",
    "javier camel": "Raleigh", "marc james": "Raleigh",
    # Richmond
    "zachary gonzalez": "Richmond",
    # San Antonio
    "erick guzman juarez": "San Antonio",
    # Savannah
    "carlos ramirez": "Savannah", "dustin earls": "Savannah",
    "lance novak": "Savannah",
    # St. Louis
    "demarkco cody": "St. Louis", "derious nichols": "St. Louis",
    "jacob rhodes": "St. Louis", "tony nguyen": "St. Louis",
    "troy willis": "St. Louis",
    # Tallahassee
    "hunter burch": "Tallahassee", "matthew mccrary": "Tallahassee",
    # Wilmington
    "brye'shawn butler": "Wilmington", "jordan davis": "Wilmington",
    "josh malloch": "Wilmington", "justin alverez": "Wilmington",
    # Asheville
    "jordan koetsier": "Asheville",
}

from collections import Counter

for sheet in wb.sheetnames:
    if sheet in SKIP:
        continue
    ws = wb[sheet]
    cnt = Counter()
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is None:
                continue
            s = str(cell).strip().lower()
            if s in LOOKUP:
                cnt[LOOKUP[s]] += 1
    top = cnt.most_common(3)
    print(f"{sheet:6}  ->  {top}")
