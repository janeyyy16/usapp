"""Quick inspector for the inventory workbook so we know what columns the
Truck Stock importer should map to."""
import openpyxl
from openpyxl.utils import get_column_letter

PATH = r"C:\Users\user\Downloads\PROJECT A USAPP ALL PARTS INVENTORY USE.xlsx"

wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
print("Sheets:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n=== {name} ===  rows={ws.max_row}  cols={ws.max_column}")
    # Show the first 3 rows so we can see headers + a couple data rows
    for ridx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if ridx > 4:
            break
        print(f"  R{ridx}:", [str(v)[:40] if v is not None else "" for v in row[:20]])
